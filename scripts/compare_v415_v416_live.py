#!/usr/bin/env python3
"""
V4.15 vs V4.16 Live-Simulation Comparison
==========================================

3-way comparison:
  1. **v4.15 live** -- from the XLSX reference export
  2. **v4.15 sim**  -- offline replay using V414SignalGenerator
  3. **v4.16 sim**  -- offline replay using V416SignalGenerator

Outputs (all saved to ``reports/v4.16/sim_feb15/``):
  - Multi-sheet Excel workbook  (``v4_16_sim_trade_report.xlsx``)
  - Equity-curve overlay PNG    (``equity_comparison.png``)
  - Exit-reason analysis PNG    (``exit_reason_analysis.png``)
  - Comparison CSV              (``comparison_report.csv``)
  - Console narrative summary

Usage:
    python scripts/compare_v415_v416_live.py

    # Custom paths:
    python scripts/compare_v415_v416_live.py \\
        --v415-sim  reports/v4.16/sim_feb15/results_v4_15_sim_*.csv \\
        --v416-sim  reports/v4.16/sim_feb15/results_v4_16_sim_*.csv \\
        --xlsx      reports/v4.15/trade_history_2026-02-15.xlsx
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import pandas as pd

# ── Project paths ──────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent

# Defaults for the Feb 15 window
SIM_DIR = ROOT / "reports" / "v4.16" / "sim_feb15"
DEFAULT_TAG = "2026-02-15_00-00-00_to_2026-02-15_20-00-00"
DEFAULT_V415_SIM = SIM_DIR / f"results_v4_15_sim_{DEFAULT_TAG}.csv"
DEFAULT_V416_SIM = SIM_DIR / f"results_v4_16_sim_{DEFAULT_TAG}.csv"
DEFAULT_EQUITY = SIM_DIR / f"equity_curves_sim_{DEFAULT_TAG}.csv"
DEFAULT_XLSX = ROOT / "reports" / "v4.15" / "trade_history_2026-02-15.xlsx"


# ══════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════

def ts_to_utc(ts) -> datetime:
    """Unix epoch seconds -> UTC datetime."""
    return datetime.fromtimestamp(int(ts), tz=timezone.utc)

def fmt_ts(ts) -> str:
    """Unix epoch -> 'YYYY-MM-DD HH:MM:SS'."""
    return ts_to_utc(ts).strftime("%Y-%m-%d %H:%M:%S")


def load_sim_csv(path: Path) -> pd.DataFrame:
    """Load a simulation trade CSV."""
    df = pd.read_csv(path)
    for col in ["entry_price", "exit_price", "pnl_pct", "pnl_dollar",
                 "position_size_pct", "stop_loss", "take_profit",
                 "entry_probability", "entry_strength"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in ["entry_time", "exit_time", "bars_held"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    return df


def load_xlsx_trades(path: Path) -> Optional[pd.DataFrame]:
    """
    Load trades from the XLSX 'Trade Summary' sheet and map to the
    standard trade schema.
    """
    if not path.exists():
        return None

    ts = pd.read_excel(path, sheet_name="Trade Summary")
    col_map = {
        "Direction": "direction",
        "Entry Price": "entry_price",
        "Exit Price": "exit_price",
        "Entry Time": "entry_time",
        "Exit Time": "exit_time",
        "Price Chg %": "pnl_pct",
        "P&L $": "pnl_dollar",
        "Bars Held": "bars_held",
        "Size %": "position_size_pct",
        "Stop Loss": "stop_loss",
        "Take Profit": "take_profit",
        "Probability": "entry_probability",
        "Strength": "entry_strength",
        "Exit Reason": "reason",
    }
    df = ts.rename(columns=col_map)

    # Convert timestamp strings to unix epochs
    for tc in ["entry_time", "exit_time"]:
        if tc in df.columns and df[tc].dtype == object:
            df[tc] = pd.to_datetime(df[tc], utc=True).astype("int64") // 10**9

    keep = list(col_map.values()) + ["model_version"]
    df = df[[c for c in keep if c in df.columns]]
    df["model_version"] = "v4.15_live"
    return df


# ══════════════════════════════════════════════════════════════════════════
# Metrics
# ══════════════════════════════════════════════════════════════════════════

def compute_metrics(trades: pd.DataFrame, starting_balance: float = 1000.0) -> Dict:
    """Compute comprehensive trading metrics."""
    n = len(trades)
    if n == 0:
        return {"Trade Count": 0}

    pnl_pcts = trades["pnl_pct"].values.astype(float)
    pnl_dollars = trades["pnl_dollar"].values.astype(float)
    bars_held = trades["bars_held"].values.astype(float)

    is_win = pnl_dollars > 0
    wins = int(is_win.sum())
    losses = n - wins

    win_rate = wins / n * 100
    avg_pnl_pct = float(pnl_pcts.mean())
    avg_win_pct = float(pnl_pcts[is_win].mean()) if wins > 0 else 0.0
    avg_loss_pct = float(pnl_pcts[~is_win].mean()) if losses > 0 else 0.0

    gross_profit = float(pnl_dollars[is_win].sum()) if wins > 0 else 0.0
    gross_loss = abs(float(pnl_dollars[~is_win].sum())) if losses > 0 else 1e-8
    profit_factor = gross_profit / gross_loss
    rr_ratio = abs(avg_win_pct / avg_loss_pct) if avg_loss_pct != 0 else float("inf")

    # Equity curve
    balance = starting_balance
    equity = [balance]
    peak = balance
    max_dd = 0.0
    for d in pnl_dollars:
        balance += d
        equity.append(balance)
        peak = max(peak, balance)
        dd = (balance - peak) / peak
        max_dd = min(max_dd, dd)

    final_balance = balance
    total_pnl = final_balance - starting_balance

    # Sharpe & Sortino
    avg_hold = float(bars_held.mean())
    est_cycle = avg_hold + 15
    trades_per_year = 525600.0 / max(est_cycle, 1)
    mean_ret = float(pnl_pcts.mean())
    std_ret = float(pnl_pcts.std()) if n > 1 else 1e-8
    sharpe = (mean_ret / std_ret) * np.sqrt(trades_per_year) if std_ret > 1e-8 else 0.0

    downside = pnl_pcts[pnl_pcts < 0]
    downside_std = float(np.sqrt(np.mean(downside ** 2))) if len(downside) > 0 else 1e-8
    sortino = (mean_ret / downside_std) * np.sqrt(trades_per_year) if downside_std > 1e-8 else 0.0

    # VaR 95%
    sorted_pnl = np.sort(pnl_pcts)
    var_idx = max(0, int(np.floor(n * 0.05)) - 1)
    var95 = float(sorted_pnl[var_idx]) if n > 0 else 0.0

    # Streaks
    max_cw, max_cl, cw, cl = 0, 0, 0, 0
    for w in is_win:
        if w:
            cw += 1; cl = 0
            max_cw = max(max_cw, cw)
        else:
            cl += 1; cw = 0
            max_cl = max(max_cl, cl)

    # Kelly
    avg_win_d = float(pnl_dollars[is_win].mean()) if wins > 0 else 0
    avg_loss_d = abs(float(pnl_dollars[~is_win].mean())) if losses > 0 else 1e-8
    wl_ratio = avg_win_d / avg_loss_d if avg_loss_d > 0 else 0
    wr_frac = wins / n
    kelly = wr_frac - (1 - wr_frac) / wl_ratio if wl_ratio > 0 else 0

    return {
        "Trade Count": n,
        "Wins": wins,
        "Losses": losses,
        "Win Rate %": round(win_rate, 2),
        "Avg PnL %": round(avg_pnl_pct, 4),
        "Avg Win %": round(avg_win_pct, 4),
        "Avg Loss %": round(avg_loss_pct, 4),
        "R:R Ratio": round(rr_ratio, 4),
        "Profit Factor": round(profit_factor, 4),
        "Sharpe (ann.)": round(sharpe, 2),
        "Sortino (ann.)": round(sortino, 2),
        "VaR 95%": round(var95, 4),
        "Max DD %": round(max_dd * 100, 4),
        "Final Balance $": round(final_balance, 2),
        "Total PnL $": round(total_pnl, 2),
        "Avg Hold (bars)": round(avg_hold, 1),
        "Max Win Streak": max_cw,
        "Max Loss Streak": max_cl,
        "Kelly Criterion": round(kelly, 4),
    }


# ══════════════════════════════════════════════════════════════════════════
# Excel generation
# ══════════════════════════════════════════════════════════════════════════

def _build_equity_series(trades: pd.DataFrame, starting_balance: float):
    """Return (timestamps, balances) for an equity curve."""
    if trades.empty:
        return [], []
    bal = starting_balance
    ts_list = [int(trades["entry_time"].iloc[0])]
    eq_list = [bal]
    for _, row in trades.iterrows():
        bal += row["pnl_dollar"]
        ts_list.append(int(row["exit_time"]))
        eq_list.append(bal)
    return ts_list, eq_list


def generate_excel_report(
    df_v416: pd.DataFrame,
    m_v416: Dict,
    df_v415_live: Optional[pd.DataFrame],
    m_v415_live: Optional[Dict],
    df_v415_sim: Optional[pd.DataFrame],
    m_v415_sim: Optional[Dict],
    output_path: Path,
    starting_balance: float = 1000.0,
) -> None:
    """
    Generate a multi-sheet Excel workbook for v4.16 simulation results,
    mirroring the structure of the v4.15 XLSX export.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colour palette
    GOLD = "FFF0B90B"
    GREEN = "FF27AE60"
    RED = "FFE74C3C"
    HEADER_BG = "FF1A1A2E"
    HEADER_FG = "FFFFFFFF"

    def style_header(ws):
        hfill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        hfont = Font(bold=True, color=HEADER_FG, size=11)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    def auto_width(ws):
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[letter].width = min(max_len + 3, 30)

    wb = __import__("openpyxl").Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Dashboard ──────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard")
    ws.append(["V4.16 Simulation Trade Report"])
    ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC",
               "", "", "Model: v4.16"])
    ws.append([])

    # Account overview + performance
    ws.append(["ACCOUNT OVERVIEW", "", "", "PERFORMANCE", ""])
    kpis = [
        ("Starting Balance", f"${starting_balance:,.2f}",
         "Win Rate", f"{m_v416.get('Win Rate %', 0)}%"),
        ("Final Balance", f"${m_v416.get('Final Balance $', 0):,.2f}",
         "Profit Factor", f"{m_v416.get('Profit Factor', 0):.2f}"),
        ("Total PnL $", f"${m_v416.get('Total PnL $', 0):+,.2f}",
         "Max Drawdown", f"{abs(m_v416.get('Max DD %', 0)):.2f}%"),
        ("Total Trades", str(m_v416.get("Trade Count", 0)),
         "Avg Bars Held", str(m_v416.get("Avg Hold (bars)", 0))),
        ("", "",
         "Sharpe Ratio", f"{m_v416.get('Sharpe (ann.)', 0):.2f}"),
    ]
    for a, b, c, d in kpis:
        ws.append([a, b, "", c, d])

    ws.append([])
    # Health checks
    ws.append(["HEALTH CHECKS", "", "", "DIRECTION BREAKDOWN", ""])
    checks = [
        ("Win Rate > 50%", m_v416.get("Win Rate %", 0) > 50),
        ("Profit Factor > 1.0", m_v416.get("Profit Factor", 0) > 1),
        ("Max Drawdown < 5%", abs(m_v416.get("Max DD %", 0)) < 5),
        ("Sharpe > 1.0", m_v416.get("Sharpe (ann.)", 0) > 1),
        ("Positive Expectancy", m_v416.get("Total PnL $", 0) > 0),
    ]
    # Direction breakdown
    long_trades = df_v416[df_v416["direction"] == "LONG"] if not df_v416.empty else pd.DataFrame()
    short_trades = df_v416[df_v416["direction"] == "SHORT"] if not df_v416.empty else pd.DataFrame()
    dir_info = [
        ("LONG Trades", str(len(long_trades))),
        ("LONG Win Rate", f"{(long_trades['pnl_dollar'] > 0).mean()*100:.1f}%" if len(long_trades) > 0 else "-"),
        ("LONG Total PnL", f"${long_trades['pnl_dollar'].sum():.2f}" if len(long_trades) > 0 else "$0.00"),
        ("SHORT Trades", str(len(short_trades))),
        ("SHORT Win Rate", f"{(short_trades['pnl_dollar'] > 0).mean()*100:.1f}%" if len(short_trades) > 0 else "-"),
    ]
    for i in range(max(len(checks), len(dir_info))):
        c_label, c_pass = checks[i] if i < len(checks) else ("", True)
        d_label, d_val = dir_info[i] if i < len(dir_info) else ("", "")
        ws.append([c_label, "PASS" if c_pass else "FAIL", "", d_label, d_val])

    style_header(ws)
    auto_width(ws)
    # Colour the pass/fail column
    for row_num in range(12, 12 + len(checks)):
        cell = ws.cell(row=row_num, column=2)
        if cell.value == "PASS":
            cell.font = Font(bold=True, color=GREEN)
        elif cell.value == "FAIL":
            cell.font = Font(bold=True, color=RED)

    # Title styling
    ws.cell(1, 1).font = Font(bold=True, size=14, color=GOLD)
    ws.cell(4, 1).font = Font(bold=True, size=11, color=GOLD)
    ws.cell(4, 4).font = Font(bold=True, size=11, color=GOLD)
    ws.cell(11, 1).font = Font(bold=True, size=11, color=GOLD)
    ws.cell(11, 4).font = Font(bold=True, size=11, color=GOLD)

    # ── Sheet 2: Trade Summary ──────────────────────────────────────────
    ws2 = wb.create_sheet("Trade Summary")
    headers = ["#", "Direction", "Entry Time", "Exit Time", "Entry Price",
               "Exit Price", "Price Chg %", "Size %", "Balance Before",
               "P&L $", "Portfolio Rtn %", "Bars Held", "Stop Loss",
               "Take Profit", "Probability", "Strength", "Exit Reason"]
    ws2.append(headers)

    bal = starting_balance
    for i, (_, t) in enumerate(df_v416.iterrows()):
        bal_before = bal
        portfolio_rtn = (t["pnl_dollar"] / bal_before * 100) if bal_before > 0 else 0
        ws2.append([
            i + 1,
            t.get("direction", ""),
            fmt_ts(t["entry_time"]),
            fmt_ts(t["exit_time"]),
            round(t["entry_price"], 4),
            round(t["exit_price"], 4),
            round(t["pnl_pct"], 4),
            round(t.get("position_size_pct", 0), 1),
            round(bal_before, 2),
            round(t["pnl_dollar"], 2),
            round(portfolio_rtn, 4),
            int(t.get("bars_held", 0)),
            round(t.get("stop_loss", 0), 4) if pd.notna(t.get("stop_loss")) else "-",
            round(t.get("take_profit", 0), 4) if pd.notna(t.get("take_profit")) else "-",
            round(t.get("entry_probability", 0), 4) if pd.notna(t.get("entry_probability")) else "-",
            round(t.get("entry_strength", 0), 4) if pd.notna(t.get("entry_strength")) else "-",
            t.get("reason", ""),
        ])
        bal += t["pnl_dollar"]

        # Colour PnL cells
        row_num = i + 2
        pnl_color = GREEN if t["pnl_dollar"] > 0 else RED if t["pnl_dollar"] < 0 else "FF888888"
        for col_idx in [7, 10, 11]:  # Price Chg %, P&L $, Portfolio Rtn %
            ws2.cell(row=row_num, column=col_idx).font = Font(color=pnl_color)
        # Direction colour
        dir_color = GREEN if t.get("direction") == "LONG" else RED
        ws2.cell(row=row_num, column=2).font = Font(color=dir_color)

    style_header(ws2)
    auto_width(ws2)

    # ── Sheet 3: Statistics ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Statistics")
    ws3.append(["Metric", "Value"])
    stat_rows = [
        ("Starting Balance", f"${starting_balance:,.2f}"),
        ("Final Balance", f"${m_v416.get('Final Balance $', 0):,.2f}"),
        ("Total PnL $", f"${m_v416.get('Total PnL $', 0):+,.2f}"),
        ("Total Return %", f"{m_v416.get('Total PnL $', 0)/starting_balance*100:.2f}"),
        ("", ""),
        ("Total Trades", str(m_v416.get("Trade Count", 0))),
        ("Wins", str(m_v416.get("Wins", 0))),
        ("Losses", str(m_v416.get("Losses", 0))),
        ("Win Rate %", f"{m_v416.get('Win Rate %', 0)}"),
        ("Profit Factor", f"{m_v416.get('Profit Factor', 0):.4f}"),
        ("", ""),
        ("Avg PnL %", f"{m_v416.get('Avg PnL %', 0):.4f}"),
        ("Avg Win (Price Chg %)", f"{m_v416.get('Avg Win %', 0):.4f}"),
        ("Avg Loss (Price Chg %)", f"{m_v416.get('Avg Loss %', 0):.4f}"),
        ("R:R Ratio", f"{m_v416.get('R:R Ratio', 0):.4f}"),
        ("", ""),
        ("Max Drawdown %", f"{abs(m_v416.get('Max DD %', 0)):.4f}"),
        ("Avg Bars Held", f"{m_v416.get('Avg Hold (bars)', 0)}"),
    ]
    # Best / worst trade
    if not df_v416.empty:
        best_pnl = df_v416["pnl_dollar"].max()
        worst_pnl = df_v416["pnl_dollar"].min()
        best_pct = df_v416["pnl_pct"].max()
        worst_pct = df_v416["pnl_pct"].min()
        stat_rows += [
            ("", ""),
            ("Best Trade $", f"${best_pnl:.2f}"),
            ("Worst Trade $", f"${worst_pnl:.2f}"),
            ("Best Trade (Price Chg %)", f"{best_pct:.4f}"),
            ("Worst Trade (Price Chg %)", f"{worst_pct:.4f}"),
        ]

    for label, val in stat_rows:
        ws3.append([label, val])
    style_header(ws3)
    auto_width(ws3)

    # ── Sheet 4: Risk Metrics ───────────────────────────────────────────
    ws4 = wb.create_sheet("Risk Metrics")
    ws4.append(["Metric", "Value"])
    risk_rows = [
        ("Sharpe Ratio (annualized)", f"{m_v416.get('Sharpe (ann.)', 0):.2f}"),
        ("Sortino Ratio (annualized)", f"{m_v416.get('Sortino (ann.)', 0):.2f}"),
        ("Value at Risk (95%)", f"{m_v416.get('VaR 95%', 0):.4f}%"),
        ("Max Drawdown %", f"{abs(m_v416.get('Max DD %', 0)):.4f}"),
        ("", ""),
        ("Kelly Criterion", f"{m_v416.get('Kelly Criterion', 0):.4f}"),
        ("Max Consecutive Wins", str(m_v416.get("Max Win Streak", 0))),
        ("Max Consecutive Losses", str(m_v416.get("Max Loss Streak", 0))),
    ]
    for label, val in risk_rows:
        ws4.append([label, val])
    style_header(ws4)
    auto_width(ws4)

    # ── Sheet 5: Comparison ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Comparison")
    # Build header row dynamically
    comp_headers = ["Metric", "v4.16 sim"]
    datasets = [("v4.16 sim", m_v416)]
    if m_v415_live is not None:
        comp_headers.insert(1, "v4.15 live")
        datasets.insert(0, ("v4.15 live", m_v415_live))
    if m_v415_sim is not None:
        comp_headers.insert(len(comp_headers) - 1 if m_v415_live else 1, "v4.15 sim")
        datasets.insert(len(datasets) - 1 if m_v415_live else 0, ("v4.15 sim", m_v415_sim))

    ws5.append(comp_headers)
    all_keys = list(m_v416.keys())
    for key in all_keys:
        row = [key]
        for _, metrics in datasets:
            row.append(metrics.get(key, ""))
        ws5.append(row)
    style_header(ws5)
    auto_width(ws5)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Excel report saved to {output_path}")


# ══════════════════════════════════════════════════════════════════════════
# Plotting
# ══════════════════════════════════════════════════════════════════════════

def plot_equity_curves(
    datasets: List[tuple],  # [(label, df, color, linestyle)]
    output_path: Path,
    starting_balance: float = 1000.0,
) -> None:
    """Plot overlaid equity curves."""
    fig, ax = plt.subplots(figsize=(14, 6))

    for label, df, color, ls in datasets:
        if df is None or df.empty:
            continue
        bal = starting_balance
        ts_list = [pd.Timestamp(ts_to_utc(int(df["entry_time"].iloc[0])))]
        eq_list = [bal]
        for _, row in df.iterrows():
            bal += row["pnl_dollar"]
            ts_list.append(pd.Timestamp(ts_to_utc(int(row["exit_time"]))))
            eq_list.append(bal)
        ax.plot(ts_list, eq_list, color=color, linestyle=ls, linewidth=1.5,
                label=f"{label} (${bal - starting_balance:+.2f})")

    ax.set_title("Equity Curve: V4.15 Live vs V4.15 Sim vs V4.16 Sim", fontsize=13)
    ax.set_ylabel("Balance ($)", fontsize=11)
    ax.legend(fontsize=10, loc="best")
    ax.grid(True, alpha=0.3)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.set_xlabel("Time (UTC, Feb 15)", fontsize=11)
    fig.autofmt_xdate()
    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Equity plot saved to {output_path}")


def plot_exit_reason_analysis(
    datasets: List[tuple],  # [(label, df, color)]
    output_path: Path,
) -> None:
    """Bar chart: trade count and PnL by exit reason for each dataset."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    # Gather all reasons
    all_reasons = set()
    for _, df, _ in datasets:
        if df is not None and not df.empty and "reason" in df.columns:
            all_reasons.update(df["reason"].unique())
    all_reasons = sorted(all_reasons)

    if not all_reasons:
        plt.close()
        return

    x = np.arange(len(all_reasons))
    n_ds = len(datasets)
    width = 0.8 / max(n_ds, 1)

    for i, (label, df, color) in enumerate(datasets):
        if df is None or df.empty or "reason" not in df.columns:
            continue
        counts = []
        pnls = []
        for r in all_reasons:
            grp = df[df["reason"] == r]
            counts.append(len(grp))
            pnls.append(grp["pnl_dollar"].sum() if len(grp) > 0 else 0)

        offset = (i - n_ds / 2 + 0.5) * width
        axes[0].bar(x + offset, counts, width, label=label, color=color, alpha=0.7)
        axes[1].bar(x + offset, pnls, width, label=label, color=color, alpha=0.7)

    axes[0].set_title("Trade Count by Exit Reason", fontsize=12)
    axes[0].set_xticks(x)
    axes[0].set_xticklabels(all_reasons, rotation=30, ha="right")
    axes[0].legend(fontsize=9)
    axes[0].grid(True, alpha=0.3, axis="y")

    axes[1].set_title("Total PnL ($) by Exit Reason", fontsize=12)
    axes[1].set_xticks(x)
    axes[1].set_xticklabels(all_reasons, rotation=30, ha="right")
    axes[1].axhline(0, color="grey", linewidth=0.8)
    axes[1].legend(fontsize=9)
    axes[1].grid(True, alpha=0.3, axis="y")

    plt.tight_layout()
    plt.savefig(output_path, dpi=150)
    plt.close()
    print(f"  Exit reason plot saved to {output_path}")


# ══════════════════════════════════════════════════════════════════════════
# Narrative
# ══════════════════════════════════════════════════════════════════════════

def generate_narrative(m_live: Dict, m_sim416: Dict) -> str:
    """Generate a short text summary comparing v4.15 live vs v4.16 sim."""
    lines = []
    lines.append("=" * 72)
    lines.append("PERFORMANCE COMPARISON: V4.15 (live) vs V4.16 (sim)")
    lines.append("=" * 72)

    if m_live.get("Trade Count", 0) == 0 or m_sim416.get("Trade Count", 0) == 0:
        lines.append("  One or both datasets produced no trades.")
        return "\n".join(lines)

    wins_v416 = 0
    wins_v415 = 0

    comparisons = [
        ("Total PnL $", "higher", True),
        ("Max DD %", "shallower", False),
        ("Sharpe (ann.)", "higher", True),
        ("Win Rate %", "higher", True),
        ("Profit Factor", "higher", True),
        ("R:R Ratio", "higher", True),
    ]

    for metric, desc, higher_better in comparisons:
        v15 = m_live.get(metric, 0)
        v16 = m_sim416.get(metric, 0)
        # For Max DD % (negative values), closer to 0 is better
        winner = "v4.16" if v16 > v15 else "v4.15"
        if winner == "v4.16":
            wins_v416 += 1
        else:
            wins_v415 += 1
        lines.append(
            f"  {metric:20s}  v4.15_live={v15:>10}  v4.16_sim={v16:>10}  -> {winner} ({desc})"
        )

    lines.append("")
    if wins_v416 > wins_v415:
        lines.append(f"  VERDICT: V4.16 OUTPERFORMS V4.15 ({wins_v416}/{wins_v416+wins_v415} metrics)")
    elif wins_v415 > wins_v416:
        lines.append(f"  VERDICT: V4.15 OUTPERFORMS V4.16 ({wins_v415}/{wins_v416+wins_v415} metrics)")
    else:
        lines.append("  VERDICT: COMPARABLE PERFORMANCE")

    lines.append(
        f"\n  Trades: v4.15_live={m_live['Trade Count']}, v4.16_sim={m_sim416['Trade Count']}"
    )
    lines.append("=" * 72)
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="V4.15 vs V4.16 comparison")
    p.add_argument("--v415-sim", type=str, default=str(DEFAULT_V415_SIM))
    p.add_argument("--v416-sim", type=str, default=str(DEFAULT_V416_SIM))
    p.add_argument("--xlsx", type=str, default=str(DEFAULT_XLSX))
    p.add_argument("--output-dir", type=str, default=str(SIM_DIR))
    p.add_argument("--balance", type=float, default=1000.0)
    return p.parse_args()


def main():
    args = parse_args()
    out = Path(args.output_dir)
    out.mkdir(parents=True, exist_ok=True)

    print("=" * 72)
    print("V4.15 vs V4.16 COMPARISON (3-way)")
    print("=" * 72)

    # ── Load data ───────────────────────────────────────────────────────
    # v4.15 live (XLSX) — primary baseline
    xlsx_path = Path(args.xlsx)
    df_v415_live = None
    m_v415_live = None
    if xlsx_path.exists():
        df_v415_live = load_xlsx_trades(xlsx_path)
        if df_v415_live is not None and not df_v415_live.empty:
            m_v415_live = compute_metrics(df_v415_live, starting_balance=args.balance)
            print(f"  v4.15 live (XLSX): {len(df_v415_live)} trades from {xlsx_path.name}")
    else:
        print(f"  XLSX not found: {xlsx_path}")

    # v4.15 sim
    v415_sim_path = Path(args.v415_sim)
    df_v415_sim = None
    m_v415_sim = None
    if v415_sim_path.exists():
        df_v415_sim = load_sim_csv(v415_sim_path)
        m_v415_sim = compute_metrics(df_v415_sim, starting_balance=args.balance)
        print(f"  v4.15 sim:        {len(df_v415_sim)} trades from {v415_sim_path.name}")
    else:
        print(f"  v4.15 sim CSV not found: {v415_sim_path}")

    # v4.16 sim
    v416_sim_path = Path(args.v416_sim)
    if not v416_sim_path.exists():
        print(f"ERROR: v4.16 sim CSV not found: {v416_sim_path}")
        sys.exit(1)
    df_v416_sim = load_sim_csv(v416_sim_path)
    m_v416_sim = compute_metrics(df_v416_sim, starting_balance=args.balance)
    print(f"  v4.16 sim:        {len(df_v416_sim)} trades from {v416_sim_path.name}")

    # ── Metrics comparison table ────────────────────────────────────────
    print("\n" + "-" * 72)
    print("METRICS COMPARISON")
    print("-" * 72)

    # Build header
    col_labels = []
    col_data = []
    if m_v415_live:
        col_labels.append("v4.15 live")
        col_data.append(m_v415_live)
    if m_v415_sim:
        col_labels.append("v4.15 sim")
        col_data.append(m_v415_sim)
    col_labels.append("v4.16 sim")
    col_data.append(m_v416_sim)

    header = f"  {'Metric':25s}" + "".join(f"  {l:>12s}" for l in col_labels)
    print(header)
    print(f"  {'─'*25}" + "  ".join(f"{'─'*12}" for _ in col_labels))

    metrics_table = []
    for key in m_v416_sim.keys():
        row_str = f"  {key:25s}"
        row_dict = {"metric": key}
        for label, metrics in zip(col_labels, col_data):
            val = metrics.get(key, "")
            row_str += f"  {str(val):>12s}"
            row_dict[label] = val
        print(row_str)
        metrics_table.append(row_dict)

    # ── Per-direction breakdown ─────────────────────────────────────────
    print("\n" + "-" * 72)
    print("PER-DIRECTION BREAKDOWN")
    print("-" * 72)

    for label, df in [("v4.15 live", df_v415_live), ("v4.15 sim", df_v415_sim),
                       ("v4.16 sim", df_v416_sim)]:
        if df is None or df.empty:
            continue
        print(f"\n  {label}:")
        for direction, grp in df.groupby("direction"):
            pnl = grp["pnl_dollar"].values
            wins = int((pnl > 0).sum())
            n = len(grp)
            wr = wins / n * 100 if n > 0 else 0
            print(f"    {direction:6s}  Trades={n:3d}  W={wins:3d}  WR={wr:.1f}%  "
                  f"PnL=${pnl.sum():.2f}")

    # ── Exit reason analysis ────────────────────────────────────────────
    print("\n" + "-" * 72)
    print("EXIT REASON ANALYSIS")
    print("-" * 72)

    for label, df in [("v4.15 live", df_v415_live), ("v4.15 sim", df_v415_sim),
                       ("v4.16 sim", df_v416_sim)]:
        if df is None or df.empty or "reason" not in df.columns:
            continue
        reason_grp = df.groupby("reason").agg(
            Trades=("pnl_dollar", "count"),
            Total_PnL=("pnl_dollar", "sum"),
            Avg_PnL=("pnl_dollar", "mean"),
            Win_Rate=("pnl_dollar", lambda x: (x > 0).mean() * 100),
        ).round(2)
        print(f"\n  {label}:")
        print(reason_grp.to_string())

    # ── Generate images ─────────────────────────────────────────────────
    print("\n" + "-" * 72)
    print("GENERATING OUTPUTS")
    print("-" * 72)

    # Equity curves
    eq_datasets = []
    if df_v415_live is not None and not df_v415_live.empty:
        eq_datasets.append(("v4.15 live", df_v415_live, "#2ecc71", "--"))
    if df_v415_sim is not None and not df_v415_sim.empty:
        eq_datasets.append(("v4.15 sim", df_v415_sim, "#3498db", "-"))
    eq_datasets.append(("v4.16 sim", df_v416_sim, "#e74c3c", "-"))

    plot_equity_curves(eq_datasets, out / "equity_comparison.png",
                       starting_balance=args.balance)

    # Exit reason analysis
    reason_datasets = []
    if df_v415_live is not None and not df_v415_live.empty:
        reason_datasets.append(("v4.15 live", df_v415_live, "#2ecc71"))
    if df_v415_sim is not None and not df_v415_sim.empty:
        reason_datasets.append(("v4.15 sim", df_v415_sim, "#3498db"))
    reason_datasets.append(("v4.16 sim", df_v416_sim, "#e74c3c"))

    plot_exit_reason_analysis(reason_datasets, out / "exit_reason_analysis.png")

    # Excel report
    generate_excel_report(
        df_v416=df_v416_sim,
        m_v416=m_v416_sim,
        df_v415_live=df_v415_live,
        m_v415_live=m_v415_live,
        df_v415_sim=df_v415_sim,
        m_v415_sim=m_v415_sim,
        output_path=out / "v4_16_sim_trade_report.xlsx",
        starting_balance=args.balance,
    )

    # Comparison CSV
    report_path = out / "comparison_report.csv"
    pd.DataFrame(metrics_table).to_csv(report_path, index=False)
    print(f"  Comparison CSV saved to {report_path}")

    # ── Narrative ───────────────────────────────────────────────────────
    if m_v415_live:
        narrative = generate_narrative(m_v415_live, m_v416_sim)
        print("\n" + narrative)

        # Also compare v4.15 sim fidelity
        if m_v415_sim:
            print("\n" + "-" * 72)
            print("SIMULATION FIDELITY: v4.15 live vs v4.15 sim")
            print("-" * 72)
            for key in ["Trade Count", "Win Rate %", "Total PnL $", "Profit Factor",
                         "Max DD %", "Avg Hold (bars)"]:
                live_val = m_v415_live.get(key, "N/A")
                sim_val = m_v415_sim.get(key, "N/A")
                print(f"  {key:25s}  live={str(live_val):>12s}  sim={str(sim_val):>12s}")

    print(f"\n  All outputs saved to {out}/")
    print("  Done!")


if __name__ == "__main__":
    main()
